#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
名单查阅系统 — 从Excel自动生成加密HTML
读取名单Excel，加密数据，生成可部署的名单查阅.html + 名单查阅.json
"""

import argparse
import base64
import hashlib
import io
import json as _json
import logging
import os
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

import openpyxl
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend

from PIL import Image


def compress_image(raw_data: bytes, max_size: int = 120, quality: int = 55) -> bytes:
    """压缩图片到指定最大尺寸，返回JPEG bytes"""
    img = Image.open(io.BytesIO(raw_data))
    w, h = img.size
    if w > h:
        new_w = max_size
        new_h = int(h * max_size / w)
    else:
        new_h = max_size
        new_w = int(w * max_size / h)
    img = img.resize((new_w, new_h), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality, optimize=True)
    return buf.getvalue()


def get_user_password(phone: str, idcard: str) -> str:
    """用户密码 = 完整身份证号（不区分大小写）"""
    return str(idcard).strip().lower()


def encrypt_aes_gcm(password: str, plaintext: str) -> str:
    """AES-256-GCM encrypt -> hex string"""
    salt = os.urandom(16)
    nonce = os.urandom(12)
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=100000,
        backend=default_backend(),
    )
    key = kdf.derive(password.encode())
    cipher = Cipher(algorithms.AES(key), modes.GCM(nonce), backend=default_backend())
    encryptor = cipher.encryptor()
    ct = encryptor.update(plaintext.encode()) + encryptor.finalize()
    tag = encryptor.tag
    return (salt + nonce + tag + ct).hex()


def sha256_first16(text: str) -> str:
    return hashlib.sha256(text.encode()).hexdigest()[:16]


def sha256_full(text: str) -> str:
    return hashlib.sha256(text.encode()).hexdigest()


def extract_qr_images(xlsx_path: str) -> dict:
    """解析 xlsx 中 DISPIMG 嵌入的图片，返回 {图片名: (raw_bytes, extension)}"""
    result = {}
    with zipfile.ZipFile(xlsx_path) as z:
        name_to_rid = {}
        try:
            ci_xml = z.read("xl/cellimages.xml").decode("utf-8")
            root = ET.fromstring(ci_xml)
            for ci in root.iter("{http://www.wps.cn/officeDocument/2017/etCustomData}cellImage"):
                pic = ci.find(".//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}pic")
                if pic is not None:
                    name_el = pic.find(".//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}cNvPr")
                    blip = pic.find(".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip")
                    if name_el is not None and blip is not None:
                        img_name = name_el.get("name", "")
                        rid = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed", "")
                        if img_name and rid:
                            name_to_rid[img_name] = rid
        except (KeyError, ET.ParseError):
            pass

        rid_to_media = {}
        try:
            rels_xml = z.read("xl/_rels/cellimages.xml.rels").decode("utf-8")
            rels_root = ET.fromstring(rels_xml)
            for rel in rels_root:
                rid = rel.get("Id", "")
                target = rel.get("Target", "")
                if rid and target:
                    rid_to_media[rid] = target
        except (KeyError, ET.ParseError):
            pass

        media_data = {}
        for fname in z.namelist():
            if fname.startswith("xl/media/") and not fname.endswith("/"):
                basename = fname.split("/")[-1]
                ext = os.path.splitext(basename)[1].lower()
                media_data[basename] = (z.read(fname), ext)

        for img_name, rid in name_to_rid.items():
            media_rel = rid_to_media.get(rid, "")
            media_file = media_rel.split("/")[-1]
            if media_file in media_data:
                result[img_name] = media_data[media_file]

    return result


def parse_dispimg(formula: str) -> str:
    """从 DISPIMG 公式中提取图片名"""
    m = re.search(r'DISPIMG\s*\(\s*"([^"]+)"', formula or "")
    return m.group(1) if m else ""


def build_rows(xlsx_path, admin_password):
    """读取Excel -> 返回加密后的数据列表"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    qr_images = extract_qr_images(xlsx_path)
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, name, phone, idcard, qr_formula, referrer = row[:6]
        if seq is None and name is None:
            continue
        if seq is not None:
            seq = int(seq)
        phone = str(phone).strip() if phone is not None else ""
        idcard = str(idcard).strip().upper() if idcard is not None else ""
        referrer = str(referrer).strip() if referrer is not None else ""
        name = str(name).strip() if name is not None else ""
        if not phone or not idcard:
            continue

        img_name = parse_dispimg(qr_formula)
        img_data = qr_images.get(img_name)
        if img_data:
            raw_bytes, _ext = img_data
            compressed = compress_image(raw_bytes, max_size=120, quality=55)
            qr_b64 = base64.b64encode(compressed).decode()
        else:
            qr_b64 = ""

        user_pwd = get_user_password(phone, idcard)
        rows.append({
            "\u5e8f\u53f7": seq,
            "\u59d3\u540d": name,
            "\u7535\u8bdd_enc": encrypt_aes_gcm(user_pwd, phone),
            "\u8eab\u4efd\u8bc1_enc": encrypt_aes_gcm(user_pwd, idcard),
            "\u7535\u8bdd_admin": encrypt_aes_gcm(admin_password, phone),
            "\u8eab\u4efd\u8bc1_admin": encrypt_aes_gcm(admin_password, idcard),
            "\u5bc6\u7801_hash": sha256_first16(user_pwd),
            "\u63a8\u8350\u4eba": referrer,
            "\u4e8c\u7ef4\u7801": qr_b64,
        })
    return rows


def generate_html(xlsx_path, admin_password, output_dir="."):
    """生成 名单查阅.html + 名单查阅.json"""
    rows = build_rows(xlsx_path, admin_password)

    # 数据转JSON字符串 → base64嵌入
    json_str = _json.dumps(rows, ensure_ascii=False, separators=(",", ":"))
    data_b64 = base64.b64encode(json_str.encode("utf-8")).decode()

    admin_hash = sha256_full(admin_password)

    css = (
        "*{margin:0;padding:0;box-sizing:border-box}"
        "body{font-family:-apple-system,BlinkMacSystemFont,Segoe UI,sans-serif;background:#f0f2f5;min-height:100dvh;display:flex;flex-direction:column}"
        ".container{background:#fff;box-shadow:0 2px 12px rgba(0,0,0,.1);width:100%;max-width:100%;min-height:100dvh;display:flex;flex-direction:column}"
        ".header{background:linear-gradient(135deg,#1677ff,#0958d9);color:#fff;padding:24px 32px;text-align:center}"
        ".header h1{font-size:22px;font-weight:600}.header p{font-size:13px;opacity:.8;margin-top:4px}"
        ".input-area{padding:16px;display:flex;gap:10px;border-bottom:1px solid #f0f0f0;align-items:center;flex-wrap:wrap}"
        ".input-area input{flex:1;min-width:160px;padding:12px 14px;border:1px solid #d9d9d9;border-radius:8px;font-size:16px;outline:none}"
        ".input-area input:focus{border-color:#1677ff;box-shadow:0 0 0 2px rgba(22,119,255,.1)}"
        ".btn-primary{padding:10px 24px;background:#1677ff;color:#fff;border:none;border-radius:8px;font-size:16px;cursor:pointer}"
        ".btn-primary:hover{background:#0958d9}.btn-outline{background:none;border:1px solid #d9d9d9;color:#666;padding:8px 16px;border-radius:8px;cursor:pointer;font-size:13px}"
        ".btn-danger{background:none;border:1px solid #ffccc7;color:#ff4d4f;padding:8px 16px;border-radius:8px;cursor:pointer;font-size:13px}"
        ".info-bar{padding:12px 32px;background:#e6f4ff;border-bottom:1px solid #91caff;color:#1677ff;font-size:13px;display:none}"
        ".info-bar.show{display:block}.info-bar.admin-bar{background:#fffbe6;border-color:#ffe58f;color:#ad6800}"
        ".table-wrap{padding:16px;overflow-x:auto}"
        "@media(max-width:600px){.header{padding:18px 16px}.header h1{font-size:18px}.table-wrap{padding:8px}th,td{padding:8px}.btn-primary{padding:12px 20px}.info-bar{padding:10px 16px}}"
        "table{width:100%;border-collapse:collapse;font-size:14px}"
        "th{background:#fafafa;padding:12px 16px;text-align:left;font-weight:600;color:#333;border-bottom:2px solid #f0f0f0}"
        "td{padding:12px 16px;border-bottom:1px solid #f0f0f0;color:#555}tr:hover td{background:#f5f5f5}"
        ".visible-cell{font-family:SF Mono,Monaco,monospace;font-size:13px;white-space:nowrap}"
        ".qr-placeholder{color:#bbb;font-style:italic;font-size:12px}"
        ".empty{text-align:center;padding:40px;color:#999;font-size:15px}"
        ".error{background:#fff2f0;color:#ff4d4f;padding:10px 16px;margin:12px 32px;border-radius:8px;border:1px solid #ffccc7;display:none;font-size:13px}"
        ".error.show{display:block}.hint{margin-top:16px;padding:0 32px 24px;color:#999;font-size:12px;text-align:center}"
        ".qr-modal{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.8);z-index:9999;justify-content:center;align-items:center;cursor:pointer}"
        ".qr-modal.show{display:flex}.qr-modal img{max-width:90vw;max-height:90vh;border-radius:8px}"
        ".qr-img{cursor:pointer}.qr-img:hover{transform:scale(1.05)}"
        ".loading{display:none;text-align:center;padding:40px;color:#999}"
        ".loading.show{display:block}.loading .spinner{display:inline-block;width:32px;height:32px;border:3px solid #e8e8e8;border-top-color:#1677ff;border-radius:50%;animation:spin .8s linear infinite;margin-bottom:12px}"
        "@keyframes spin{to{transform:rotate(360deg)}}"
        ".pager{display:none;padding:12px 16px;border-top:1px solid #f0f0f0;text-align:center;gap:8px;align-items:center;justify-content:center;flex-wrap:wrap}"
        ".pager.show{display:flex}.pager .page-btn{padding:6px 14px;border:1px solid #d9d9d9;border-radius:6px;background:#fff;cursor:pointer;font-size:13px}"
        ".pager .page-btn:disabled{opacity:.4;cursor:default}.pager .page-info{font-size:13px;color:#999}"
        "@media(max-width:600px){.pager{padding:10px 8px}.pager .page-btn{padding:8px 12px;font-size:12px}}"
    )

    js = (
        'var A="' + admin_hash + '",B="' + data_b64 + '",N=100,R=[],C=[],CP=0,M=0;'
        'function $(i){return document.getElementById(i);}'
        'function S(i){$(i).classList.add("show");}'
        'function H(i){$(i).classList.remove("show");}'
        'function E(m){var e=$("errorMsg");e.textContent=m;e.classList.add("show");setTimeout(function(){e.classList.remove("show")},7000);}'
        'try{var _b=atob(B);var _u=new Uint8Array(_b.length);for(var _i=0;_i<_b.length;_i++){_u[_i]=_b.charCodeAt(_i);}R=JSON.parse(new TextDecoder("utf-8").decode(_u));}catch(e){E("数据解析失败");}'
        'function hb(h){return new Uint8Array(h.match(/.{1,2}/g).map(function(b){return parseInt(b,16)})).buffer;}'
        'function bh(b){return Array.from(new Uint8Array(b)).map(function(x){return x.toString(16).padStart(2,"0")}).join("");}'
        'async function s256(s){var b=new TextEncoder().encode(s);var h=await crypto.subtle.digest("SHA-256",b);return bh(h);}'
        'async function s16(s){return(await s256(s)).substring(0,16);}'
        'async function dc(e,p){'
        'var r=hb(e),s=r.slice(0,16),n=r.slice(16,28),t=r.slice(28,44),c=r.slice(44),km=new TextEncoder().encode(p);'
        'var bk=await crypto.subtle.importKey("raw",km,"PBKDF2",false,["deriveKey"]);'
        'var k=await crypto.subtle.deriveKey({name:"PBKDF2",salt:s,iterations:100000,hash:"SHA-256"},bk,{name:"AES-GCM",length:256},false,["decrypt"]);'
        'var x=new Uint8Array(c.byteLength+t.byteLength);x.set(new Uint8Array(c),0);x.set(new Uint8Array(t),c.byteLength);'
        'try{return new TextDecoder().decode(await crypto.subtle.decrypt({name:"AES-GCM",iv:n,tagLength:128},k,x));}catch(e){return null;}}'
        'async function go(){'
        'var p=$("pwdInput").value.trim();if(!p){E("请输入密码");return;}H("errorMsg");'
        'if(!crypto.subtle){E("浏览器不支持加密API");return;}'
        'var h;try{h=await s256(p);}catch(e){E(e.message);return;}'
        'M=(h===A);var k=M?p:p.toLowerCase();C=[];'
        'if(!M){'
        'var ph=await s16(k);'
        'for(var i=0;i<R.length;i++){'
        'var w=R[i];'
        'if(w["\u5bc6\u7801_hash"]===ph){'
        'try{var t=await dc(w["\u7535\u8bdd_enc"],k);if(t){var d=await dc(w["\u8eab\u4efd\u8bc1_enc"],k);if(d){var o={};for(var z in w){o[z]=w[z];}o["\u7535\u8bdd"]=t;o["\u8eab\u4efd\u8bc1"]=d;C.push(o);}}}'
        'catch(e){}}}}'
        'else{'
        'S("loadingArea");var dn=0;'
        'for(var i=0;i<R.length;i+=50){'
        'var rs=await Promise.all(R.slice(i,i+50).map(async function(w){'
        'try{var t=await dc(w["\u7535\u8bdd_admin"],k);var d=await dc(w["\u8eab\u4efd\u8bc1_admin"],k);if(t&&d){var o={};for(var z in w){o[z]=w[z];}o["\u7535\u8bdd"]=t;o["\u8eab\u4efd\u8bc1"]=d;return o;}}'
        'catch(e){}return null;}));'
        'for(var j=0;j<rs.length;j++){if(rs[j]){C.push(rs[j]);}}'
        'dn=Math.min(i+50,R.length);$("loadingText").textContent="\u89e3\u5bc6\u4e2d "+dn+"/"+R.length;'
        'await new Promise(function(r){setTimeout(r,0);});}'
        'H("loadingArea");}'
        'if(C.length===0){E("密码错误");return;}'
        'CP=0;render();UI();}'
        'function render(){'
        'var st=CP*N,tp=Math.ceil(C.length/N);'
        'var pd=C.slice(st,st+N);'
        'if(pd.length===0){$("tableArea").innerHTML="<div class=\\\"empty\\\">\u65e0\u6570\u636e</div>";return;}'
        'var h="<table><thead><tr><th>\u5e8f\u53f7</th><th>\u59d3\u540d</th><th>\u7535\u8bdd</th><th>\u8eab\u4efd\u8bc1</th><th>\u4e8c\u7ef4\u7801</th><th>\u63a8\u8350\u4eba</th></tr></thead><tbody>";'
        'for(var i=0;i<pd.length;i++){var r=pd[i];'
        'h+="<tr><td>"+r["\u5e8f\u53f7"]+"</td><td><strong>"+r["\u59d3\u540d"]+"</strong></td>";'
        'h+="<td><span class=\\\"visible-cell\\\">"+r["\u7535\u8bdd"]+"</span></td><td><span class=\\\"visible-cell\\\">"+r["\u8eab\u4efd\u8bc1"]+"</span></td>";'
        'var q=r["\u4e8c\u7ef4\u7801"]||"";'
        'if(q){h+="<td><img class=\\\"qr-img\\\" src=\\\"data:image/jpeg;base64,"+q+"\\\" onclick=\\\"QR(this.src)\\\"></td>";}'
        'else{h+="<td><span class=\\\"qr-placeholder\\\">\u65e0</span></td>";}'
        'h+="<td>"+(r["\u63a8\u8350\u4eba"]||"-")+"</td></tr>";}'
        'h+="</tbody></table>";$("tableArea").innerHTML=h;'
        'if(tp>1){S("pagerArea");$("pageInfo").textContent="\u7b2c "+(CP+1)+"/"+tp+" \u9875(\u5171"+C.length+"\u6761)";$("pagePrev").disabled=(CP===0);$("pageNext").disabled=(CP>=tp-1);}'
        'else{H("pagerArea");}}'
        'function QR(s){$("qrModalImg").src=s;$("qrModal").classList.add("show");}'
        'function GP(d){var tp=Math.ceil(C.length/N),np=CP+d;if(np<0||np>=tp)return;CP=np;render();$("tableArea").scrollIntoView({behavior:"smooth",block:"start"});}'
        'function UI(){'
        'var b=$("infoBar");b.classList.add("show");'
        'if(M){b.classList.add("admin-bar");b.innerHTML="\u7ba1\u7406\u5458\u6a21\u5f0f - \u5168\u90e8"+C.length+"\u6761 <button class=\\\"btn-outline\\\" onclick=\\\"out()\\\">\u9000\u51fa</button>";}'
        'else{b.classList.remove("admin-bar");b.innerHTML="\u5df2\u89e3\u9501 <b>"+C[0]["\u59d3\u540d"]+"</b> <button class=\\\"btn-danger\\\" onclick=\\\"out()\\\">\u9000\u51fa</button>";}'
        '$("inputArea").style.display="none";}'
        'function out(){'
        '$("pwdInput").value="";$("inputArea").style.display="flex";'
        '$("infoBar").classList.remove("show","admin-bar");H("errorMsg");H("pagerArea");H("loadingArea");'
        '$("tableArea").innerHTML="<div class=\\\"empty\\\">\u8bf7\u8f93\u5165\u5bc6\u7801\u67e5\u770b\u6570\u636e</div>";$("pwdInput").focus();C=[];}'
        '$("pwdInput").addEventListener("keydown",function(e){if(e.key==="Enter"){go();}});'
    )

    html = (
        '<!DOCTYPE html>\n<html lang="zh-CN">\n<head>\n<meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
        '<title>\u540d\u5355\u67e5\u9605</title>\n<style>' + css + '</style>\n</head>\n<body>\n'
        '<div class="container">\n'
        '<div class="header"><h1>\u9e6c\u6218\u961f\u53a6\u95e8VS\u5e73\u6f6d\u4e8c\u7ef4\u7801\u67e5\u8be2</h1><p>\u8f93\u5165\u60a8\u7684\u8eab\u4efd\u8bc1\u53f7\u7801\u67e5\u770b\u4e2a\u4eba\u4fe1\u606f</p></div>\n'
        '<div class="input-area" id="inputArea">\n'
        '<input type="text" id="pwdInput" placeholder="\u8bf7\u8f93\u5165\u8eab\u4efd\u8bc1\u53f7\u7801" maxlength="18" autofocus>\n'
        '<button class="btn-primary" onclick="go()">\u67e5\u8be2</button>\n'
        '</div>\n'
        '<div class="info-bar" id="infoBar"></div>\n'
        '<div class="error" id="errorMsg"></div>\n'
        '<div class="table-wrap" id="tableArea"><div class="empty">\u8bf7\u8f93\u5165\u5bc6\u7801\u67e5\u770b\u6570\u636e</div></div>\n'
        '<div class="loading" id="loadingArea"><div class="spinner"></div><div id="loadingText">\u89e3\u5bc6\u4e2d...</div></div>\n'
        '<div class="pager" id="pagerArea">\n'
        '<button class="page-btn" id="pagePrev" onclick="GP(-1)">\u4e0a\u4e00\u9875</button>\n'
        '<span class="page-info" id="pageInfo">\u7b2c 1/N \u9875</span>\n'
        '<button class="page-btn" id="pageNext" onclick="GP(1)">\u4e0b\u4e00\u9875</button>\n'
        '</div>\n'
        '<div class="qr-modal" id="qrModal" onclick="this.classList.remove(\'show\')"><img src="" id="qrModalImg"></div>\n'
        '<div class="hint">\u5bc6\u7801 = \u60a8\u7684\u8eab\u4efd\u8bc1\u53f7\u7801\uff08\u4e0d\u533a\u5206\u5927\u5c0f\u5199\uff09</div>\n'
        '</div>\n'
        '<script>' + js + '</script>\n</body>\n</html>'
    )

    return html, data_b64


def setup_logging():
    log_file = os.path.join(os.getcwd(), "\u540d\u5355\u67e5\u9605\u751f\u6210\u5668.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )
    return logging.getLogger(__name__)


def main():
    logger = setup_logging()
    parser = argparse.ArgumentParser(description="\u540d\u5355\u67e5\u9605\u751f\u6210\u5668")
    parser.add_argument("xlsx", nargs="?", default="\u540d\u5355.xlsx")
    parser.add_argument("-o", "--output", default="\u540d\u5355\u67e5\u9605.html")
    parser.add_argument("--admin-pwd", default="admin888")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        logger.error("\u6587\u4ef6\u4e0d\u5b58\u5728: %s", args.xlsx)
        input("\u6309\u56de\u8f66\u952e\u9000\u51fa...")
        sys.exit(1)

    logger.info("\u8bfb\u53d6: %s", args.xlsx)
    try:
        out_dir = os.path.dirname(args.output) or "."
        html, data_b64 = generate_html(args.xlsx, args.admin_pwd, output_dir=out_dir)
    except Exception as e:
        logger.exception("\u751f\u6210\u5931\u8d25: %s", e)
        input("\u6309\u56de\u8f66\u952e\u9000\u51fa...")
        sys.exit(1)

    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)

    logger.info("\u751f\u6210\u5b8c\u6210: %s", args.output)
    # count records from embedded data
    import json as _j2
    logger.info("\u8bb0\u5f55\u6570: %d", len(_j2.loads(base64.b64decode(data_b64).decode())))
    logger.info("\u7ba1\u7406\u5458\u5bc6\u7801: %s", args.admin_pwd)
    logger.info("\u7ba1\u7406\u5458Hash: %s...", sha256_full(args.admin_pwd)[:16])


if __name__ == "__main__":
    main()
